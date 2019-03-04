from __future__ import division, print_function
import openpyxl


class ExcelReader: # https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
    def __init__(self):
        self.book = None
        self.excel_file = None

    def open_excel_file(self, excel_file):
        self.excel_file = excel_file
        self.book = openpyxl.load_workbook(self.excel_file)

    def read_workbook(self):
        dict_list = []

        for sheet in self.book.worksheets:
            # read header values into the list
            keys = [sheet.cell(1, col_index).value for col_index in range(1, sheet.max_column + 1)]

            for row_index in range(2, sheet.max_row + 1):
                d = {}
                for col_index in range(1, sheet.max_column + 1):
                    d[keys[col_index - 1]] = sheet.cell(row_index, col_index).value
                dict_list.append(d)
        return dict_list

    def __get_sheet(self, sheet_name):
        return self.book[sheet_name] if sheet_name is not None else self.book.active

    def read_value_from_cell(self, row, column, sheet_name=None):
        """ row and column indexes are 1-based """
        return self.__get_sheet(sheet_name).cell(row, column).value

    def write_value_to_cell(self, row, column, value, sheet_name=None):
        """ row and column indexes are 1-based """
        self.__get_sheet(sheet_name).cell(row, column).value = value

    def save_workbook(self, path):
        self.book.save(path)

    def append_row(self, row, sheet_name=None):
        """ appends row at the end """
        self.__get_sheet(sheet_name).append(row)

    def append_rows(self, rows, sheet_name=None):
        """ appends rows at the end """
        sheet = self.__get_sheet(sheet_name)
        for row in rows:
            sheet.append(row)

    def create_new_workbook(self):
        self.book = openpyxl.Workbook()


if __name__ == '__main__':
    reader = ExcelReader()
    # -------------- open and read --------------
    reader.open_excel_file('E:/Doc.xlsx') # xlsx only
    rows = reader.read_workbook()
    print(rows)


    # -------------- create add save --------------
    # reader.create_new_workbook()
    # reader.append_rows([['1', '2', '3'], ['4', '5', '6'], ['7', '8', '9']])
    # reader.save_workbook('E:/Doc3.xls')

